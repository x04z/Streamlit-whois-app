import streamlit as st
from streamlit_option_menu import option_menu
import pandas as pd
import requests
import time
from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED
import socket
import struct
import ipaddress
from urllib.parse import quote
import math
import altair as alt 
import json 
import io 
import re 

# --- Excelグラフ生成用ライブラリ ---
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ページ設定
st.set_page_config(layout="wide", page_title="検索大臣", page_icon="🌐")

# ==========================================
# 🛠️ 自動モード判定ロジック (st.secrets利用)
# ==========================================
# ローカル環境では secrets.toml がなくてもエラーにならないよう try-except で処理
# Cloud側で ENV_MODE = "public" が設定されている場合のみ、機能制限モード(True)になる
IS_PUBLIC_MODE = False
try:
    if "ENV_MODE" in st.secrets and st.secrets["ENV_MODE"] == "public":
        IS_PUBLIC_MODE = True
except FileNotFoundError:
    # ローカルでsecretsファイル自体がない場合は全機能モード(False)とする
    IS_PUBLIC_MODE = False
# ==========================================

# --- 設定 ---
MODE_SETTINGS = {
    "安定性重視 (2.5秒待機/単一スレッド)": {
        "MAX_WORKERS": 1, 
        "DELAY_BETWEEN_REQUESTS": 2.5 
    },
    "速度優先 (1.4秒待機/2スレッド)": {
        "MAX_WORKERS": 2, 
        "DELAY_BETWEEN_REQUESTS": 1.4 
    }
}
IP_API_URL = "http://ip-api.com/json/{ip}?fields=status,country,countryCode,isp,org,query,message" # orgを追加
IPINFO_API_URL = "https://ipinfo.io/{ip}?token={token}" # Proモード用
RDAP_BOOTSTRAP_URL = "https://rdap.apnic.net/ip/{ip}" # RDAP用

RATE_LIMIT_WAIT_SECONDS = 120 
  
RIR_LINKS = {
    'RIPE': 'https://apps.db.ripe.net/db-web-ui/#/query?searchtext={ip}',
    'ARIN': 'https://search.arin.net/rdap/?query={ip}',
    'APNIC': 'https://wq.apnic.net/static/search.html',
    'JPNIC': 'https://www.nic.ad.jp/ja/whois/ja-gateway.html',
    'AFRINIC': 'https://www.afrinic.net/whois',
    'ICANN Whois': 'https://lookup.icann.org/',
}

SECONDARY_TOOL_BASE_LINKS = {
    'VirusTotal': 'https://www.virustotal.com/',
    'Whois.com': 'https://www.whois.com/',
    'Who.is': 'https://who.is/',
    'DomainSearch.jp': 'https://www.domainsearch.jp/',
    'Aguse': 'https://www.aguse.jp/',
    'IP2Proxy': 'https://www.ip2proxy.com/',
    'DNS Checker': 'https://dnschecker.org/',
    'DNSlytics': 'https://dnslytics.com/',
    'IP Location': 'https://iplocation.io/',
    'CP-WHOIS': 'https://doco.cph.jp/whoisweb.php',
    }

COUNTRY_CODE_TO_RIR = {
    'JP': 'JPNIC', 'CN': 'APNIC', 'AU': 'APNIC', 'KR': 'APNIC', 'IN': 'APNIC',
    'ID': 'APNIC', 'MY': 'APNIC', 'NZ': 'APNIC', 'SG': 'APNIC',
    'TH': 'APNIC', 'VN': 'APNIC', 'PH': 'APNIC', 'PK': 'APNIC', 
    'BD': 'APNIC', 'HK': 'APNIC', 'TW': 'APNIC', 'NP': 'APNIC', 'LK': 'APNIC',
    'MO': 'APNIC', 
    'US': 'ARIN', 'CA': 'ARIN',
    'ZA': 'AFRINIC', 'EG': 'AFRINIC', 'NG': 'AFRINIC',
    'KE': 'AFRINIC', 'DZ': 'AFRINIC', 'MA': 'AFRINIC', 'GH': 'AFRINIC', 
    'CM': 'AFRINIC', 'TN': 'AFRINIC', 'ET': 'AFRINIC', 'TZ': 'AFRINIC',
    'DE': 'RIPE', 'FR': 'RIPE', 'GB': 'RIPE', 'RU': 'RIPE',
    'NL': 'RIPE', 'IT': 'RIPE', 'ES': 'RIPE', 'PL': 'RIPE', 
    'TR': 'RIPE', 'UA': 'RIPE', 'SA': 'RIPE', 'IR': 'RIPE', 
    'CH': 'RIPE', 'SE': 'RIPE', 'NO': 'RIPE', 'DK': 'RIPE', 
    'BE': 'RIPE', 'AT': 'RIPE', 'GR': 'RIPE', 'PT': 'RIPE',
    'IE': 'RIPE', 'FI': 'RIPE', 'CZ': 'RIPE', 'RO': 'RIPE',
    'HU': 'RIPE', 'IL': 'RIPE', 'KZ': 'RIPE', 'BG': 'RIPE',
    'HR': 'RIPE', 'RS': 'RIPE', 'AE': 'RIPE', 'QA': 'RIPE',
}

COUNTRY_CODE_TO_NUMERIC_ISO = {
    'AF': 4, 'AL': 8, 'DZ': 12, 'AS': 16, 'AD': 20, 'AO': 24, 'AI': 660, 'AQ': 10, 'AG': 28, 'AR': 32,
    'AM': 51, 'AW': 533, 'AU': 36, 'AT': 40, 'AZ': 31, 'BS': 44, 'BH': 48, 'BD': 50, 'BB': 52, 'BY': 112,
    'BE': 56, 'BZ': 84, 'BJ': 204, 'BM': 60, 'BT': 64, 'BO': 68, 'BA': 70, 'BW': 72, 'BV': 74, 'BR': 76,
    'VG': 92, 'IO': 86, 'BN': 96, 'BG': 100, 'BF': 854, 'BI': 108, 'KH': 116, 'CM': 120, 'CA': 124, 'CV': 132,
    'KY': 136, 'CF': 140, 'TD': 148, 'CL': 152, 'CN': 156, 'CX': 162, 'CC': 166, 'CO': 170, 'KM': 174, 'CG': 178,
    'CD': 180, 'CK': 184, 'CO': 170, 'CR': 188, 'HR': 191, 'CU': 192, 'CY': 196, 'CZ': 203, 'DK': 208, 'DJ': 262, 'DM': 212,
    'DO': 214, 'EC': 218, 'EG': 818, 'SV': 222, 'GQ': 226, 'ER': 232, 'EE': 233, 'ET': 231, 'FK': 238, 'FO': 234,
    'FJ': 242, 'FI': 246, 'FR': 250, 'GF': 254, 'PF': 258, 'TF': 260, 'GA': 266, 'GM': 270, 'GE': 268, 'DE': 276,
    'GH': 288, 'GI': 292, 'GR': 300, 'GL': 304, 'GD': 308, 'GP': 312, 'GU': 316, 'GT': 320, 'GN': 324, 'GW': 624,
    'GY': 328, 'HT': 332, 'HM': 334, 'VA': 336, 'HN': 340, 'HK': 344, 'HU': 348, 'IS': 352, 'IN': 356, 'ID': 360,
    'IR': 364, 'IQ': 368, 'IE': 372, 'IL': 376, 'IT': 380, 'CI': 384, 'JM': 388, 'JP': 392, 'JO': 400, 'KZ': 398,
    'KE': 404, 'KI': 296, 'KP': 408, 'KR': 410, 'KW': 414, 'KG': 417, 'LA': 418, 'LV': 428, 'LB': 422, 'LS': 426,
    'LR': 430, 'LY': 434, 'LI': 438, 'LT': 440, 'LU': 442, 'MO': 446, 'MK': 807, 'MG': 450, 'MW': 454, 'MY': 458,
    'MV': 462, 'ML': 466, 'MT': 470, 'MH': 584, 'MQ': 474, 'MR': 478, 'MU': 480, 'YT': 175, 'MX': 484, 'FM': 583,
    'MD': 498, 'MC': 492, 'MN': 496, 'MS': 500, 'MA': 504, 'MZ': 508, 'MM': 104, 'NA': 516, 'NR': 520, 'NP': 524,
    'NL': 528, 'AN': 530, 'NC': 540, 'NZ': 554, 'NI': 558, 'NE': 562, 'NG': 566, 'NU': 570, 'NF': 574, 'MP': 580,
    'NO': 578, 'OM': 512, 'PK': 586, 'PW': 585, 'PS': 275, 'PA': 591, 'PG': 598, 'PY': 600, 'PE': 604, 'PH': 608,
    'PN': 612, 'PL': 616, 'PT': 620, 'PR': 630, 'QA': 634, 'RE': 638, 'RO': 642, 'RU': 643, 'RW': 646, 'SH': 654,
    'KN': 659, 'LC': 662, 'PM': 666, 'VC': 670, 'WS': 882, 'SM': 674, 'ST': 678, 'SA': 682, 'SN': 686, 'RS': 688,
    'SC': 690, 'SL': 694, 'SG': 702, 'SK': 703, 'SI': 705, 'SB': 90, 'SO': 706, 'ZA': 710, 'GS': 239, 'ES': 724,
    'LK': 144, 'SD': 736, 'SR': 740, 'SJ': 744, 'SZ': 748, 'SE': 752, 'CH': 756, 'SY': 760, 'TW': 158, 'TJ': 762,
    'TZ': 834, 'TH': 764, 'TL': 626, 'TG': 768, 'TK': 772, 'TO': 776, 'TT': 780, 'TN': 788, 'TR': 792, 'TM': 795,
    'TC': 796, 'TV': 798, 'UG': 800, 'UA': 804, 'AE': 784, 'GB': 826, 'US': 840, 'UM': 581, 'UY': 858, 'UZ': 860,
    'VU': 548, 'VE': 862, 'VN': 704, 'VI': 850, 'WF': 876, 'EH': 732, 'YE': 887, 'ZM': 894, 'ZW': 716
}

# --- ISP名称の日本語マッピング (企業名統一版) ---
# ここは「完全一致」で見つかるもの
ISP_JP_NAME = {
    # --- NTT Group ---
    'NTT Communications Corporation': 'NTTドコモビジネス', 
    'NTT COMMUNICATIONS CORPORATION': 'NTTドコモビジネス',
    'NTT DOCOMO BUSINESS,Inc.': 'NTTドコモビジネス',
    'NTT DOCOMO, INC.': 'NTTドコモ',
    'NTT PC Communications, Inc.': 'NTTPCコミュニケーションズ',
    
    # --- KDDI Group ---
    'Kddi Corporation': 'KDDI',
    'KDDI CORPORATION': 'KDDI',
    'DION': 'KDDI',
    'Dion': 'KDDI',
    'dion': 'KDDI',
    'Chubu Telecommunications Co., Inc.': '中部テレコミュニケーション',
    'Chubu Telecommunications Company, Inc.': '中部テレコミュニケーション',
    'Hokkaido Telecommunication Network Co., Inc.': 'HOTnet',
    'Energia Communications, Inc.': 'エネコム',
    'STNet, Inc.': 'STNet',
    'QTNet, Inc.': 'QTNet',
    'BIGLOBE Inc.': 'ビッグローブ',
    
    # --- SoftBank Group ---
    'SoftBank Corp.': 'ソフトバンク',
    'Yahoo Japan Corporation': 'LINEヤフー',
    'LY Corporation': 'LINEヤフー',
    'LINE Corporation': 'LINEヤフー',
    
    # --- Rakuten Group ---
    'Rakuten Group, Inc.': '楽天グループ',
    'Rakuten Mobile, Inc.': '楽天モバイル',
    'Rakuten Communications Corp.': '楽天コミュニケーションズ',
    
    # --- Sony Group ---
    'Sony Network Communications Inc.': 'ソニーネットワークコミュニケーションズ',
    'So-net Entertainment Corporation': 'ソニーネットワークコミュニケーションズ', 
    'So-net Corporation': 'ソニーネットワークコミュニケーションズ',
    
    # --- Major ISPs / VNEs ---
    'Internet Initiative Japan Inc.': 'IIJ',
    'NIFTY Corporation': 'ニフティ',
    'FreeBit Co., Ltd.': 'フリービット',
    'TOKAI Communications Corporation': 'TOKAIコミュニケーションズ',
    'DREAM TRAIN INTERNET INC.': 'ドリーム・トレイン・インターネット (DTI)',
    'ASAHI Net, Inc.': '朝日ネット',
    'Asahi Net': '朝日ネット',
    'Optage Inc.': 'オプテージ',
    'Jupiter Telecommunications Co., Ltd.': 'J:COM', 
    'JCOM Co., Ltd.': 'J:COM',
    'JCN': 'J:COM', 
    'SAKURA Internet Inc.': 'さくらインターネット',
    'GMO Internet, Inc.': 'GMOインターネット',
    'INTERNET MULTIFEED CO.': 'インターネットマルチフィード',
    'IDC Frontier Inc.': 'IDCフロンティア',
    
    # --- Others ---
    'ARTERIA Networks Corporation': 'アルテリア・ネットワークス',
    'UCOM Corporation': 'アルテリア・ネットワークス',
    'VECTANT Ltd.': 'アルテリア・ネットワークス',
    'KIBI Cable Television Co., Ltd.': '吉備ケーブルテレビ',
}

# --- 🆕 強力な名寄せルール (ハードコーディング辞書) ---
# 小文字・正規化されたISP名に対して、部分一致で検索し、強制変換する
ISP_REMAP_RULES = [
    # J:COM系 (古いJCN表記などを全てJ:COMへ統合)
    ('jcn', 'J:COM'),
    ('jupiter', 'J:COM'),
    ('cablenet', 'J:COM'),
    ('tsuchiura cable', 'J:COM'),
    ('kawayu', 'J:COM'), 
    
    # KDDI系
    ('dion', 'KDDI'),
    ('au one', 'KDDI (au one net)'),
    ('kddi', 'KDDI'),

    # 電力系・その他
    ('k-opti', 'オプテージ'),
    ('ctc', '中部テレコミュニケーション'),
    ('commufa', '中部テレコミュニケーション'),
    ('vectant', 'アルテリア・ネットワークス'),
    ('ucom', 'アルテリア・ネットワークス'),
    ('arteria', 'アルテリア・ネットワークス'),
    ('softbank', 'ソフトバンク'),
    ('bbtec', 'ソフトバンク'),
    ('ocn', 'OCN'),
    ('ntt', 'NTTグループ'), 
    ('infosphere', 'NTTPC (InfoSphere)'),
    ('wakwak', 'NTT-ME (WAKWAK)'),
    ('plala', 'NTTドコモ (ぷらら)'),
    ('so-net', 'ソニーネットワークコミュニケーションズ'),
    ('nuro', 'ソニーネットワークコミュニケーションズ (NURO)'),
    ('biglobe', 'ビッグローブ'),
    ('dti', 'DTI'),
    ('iij', 'IIJ'),
    ('transix', 'インターネットマルチフィード (transix)'),
    ('mfeed', 'インターネットマルチフィード'),
    ('v6plus', 'JPNE (v6プラス)'),
    ('jpne', 'JPNE'),
    ('en ne', '楽天モバイル'), 
    ('rakuten', '楽天グループ'),
]

# 正規化関数: 小文字化し、カンマ(,)とピリオド(.)を除去する
def normalize_isp_key(text):
    if not text:
        return ""
    return text.lower().replace(',', '').replace('.', '').strip()

# 検索用にキーを正規化した辞書を作成
ISP_JP_NAME_NORMALIZED = {normalize_isp_key(k): v for k, v in ISP_JP_NAME.items()}

# --- 匿名化・プロキシ判定用データ ---

@st.cache_data(ttl=86400, show_spinner=False)
def fetch_tor_exit_nodes():
    terminal = st.empty()
    log_lines = []
    
    def update_log(new_line, color="#00FF41"):
        log_lines.append(f"<span style='color:{color};'>[SYS] {new_line}</span>")
        display_text = "<br>".join(log_lines[-5:])
        terminal.markdown(f"""
            <div style="background-color: rgba(13, 2, 8, 0.9); border: 1px solid #FF0055; padding: 15px; border-radius: 8px; font-family: 'Courier New', Courier, monospace; font-size: 14px; line-height: 1.3; box-shadow: 0 0 20px rgba(255, 0, 85, 0.4); margin-bottom: 20px;">
                <div style="color: #FF0055; font-weight: bold; margin-bottom: 5px; font-size: 10px; border-bottom: 1px solid #FF0055;">ENCRYPTED DATA STREAMING...</div>
                {display_text}
            </div>
        """, unsafe_allow_html=True)
        time.sleep(0.3)

    update_log("BOOTING NEURAL LINK...")
    update_log("DECRYPTING EXIT NODE MANIFEST...")
    
    try:
        url = "https://check.torproject.org/exit-addresses"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        
        update_log("HANDSHAKE SUCCESSFUL.", "#00FFFF")
        
        exit_ips = set()
        for line in response.text.splitlines():
            if line.startswith("ExitAddress"):
                parts = line.split()
                if len(parts) >= 2:
                    exit_ips.add(parts[1])
        
        update_log(f"NODES LOADED: {len(exit_ips)} UNITS", "#00FFFF")
        update_log("SESSION SECURED. SYSTEM ONLINE.", "#00FF41")
        time.sleep(1.0)
        terminal.empty()
        return exit_ips
        
    except Exception as e:
        update_log(f"CRITICAL ERROR: {e}", "#FF0000")
        time.sleep(2.0)
        terminal.empty()
        return set()

HOSTING_VPN_KEYWORDS = [
    "hosting", "datacenter", "vps", "cloud", "server", "vpn", "proxy", "dedi",
    "amazon technologies", "amazon.com", "google llc", "google cloud", "microsoft corporation", "azure",
    "oracle cloud", "alibaba", "tencent", "huawei", "digitalocean", "linode", "vultr", "ovh", "hetzner",
    "m247", "proweb", "choopa", "leaseweb", "datacamp", "ip-volume", "flyservers", 
    "performive", "hostroyale", "packet exchange", "xtom", "tzulo", "psychz", 
    "franantech", "buyvm", "melbicom", "pfcloud", "epyc", "layerhost",
    "akamai", "cloudflare", "fastly", "cdn77", "imperva", "incapsula", "cloudfront",
    "expressvpn", "nordvpn", "proton", "mullvad", "private internet access", "windscribe",
    "cyberghost", "torguard", "vyprvpn", "purevpn"
]

def detect_proxy_vpn_tor(ip, isp_name, tor_nodes):
    isp_lower = isp_name.lower()
    if ip in tor_nodes: return "Tor Node"
    if "icloud" in isp_lower or "private relay" in isp_lower: return "iCloud Private Relay"
    privacy_keywords = ["vpn", "proxy", "applied privacy", "privacy foundation", "calyx institute", "foundation for applied privacy"]
    if any(kw in isp_lower for kw in privacy_keywords): return "VPN/Proxy (Named)"
    if any(kw in isp_lower for kw in HOSTING_VPN_KEYWORDS):
        if any(cdn in isp_lower for cdn in ["cloudflare", "akamai", "fastly", "cloudfront"]): return "CDN/Proxy"
        return "Hosting/DataCenter"
    # 修正: Residential/Business -> Standard Connection
    return "Standard Connection"

# 🆕 強化されたISP名寄せロジック
def get_jp_names(english_isp, country_code):
    if not english_isp:
        return "N/A", COUNTRY_JP_NAME.get(country_code, country_code)

    normalized_input = normalize_isp_key(english_isp)
    jp_isp = english_isp # デフォルトは英語のまま

    # 1. 完全一致 (高速)
    if english_isp in ISP_JP_NAME:
        jp_isp = ISP_JP_NAME[english_isp]
    elif normalized_input in ISP_JP_NAME_NORMALIZED:
        jp_isp = ISP_JP_NAME_NORMALIZED[normalized_input]
    else:
        # 2. 部分一致 (名寄せルール適用)
        # ルールリストの上から順にチェックし、ヒットしたら即採用
        found = False
        for keyword, mapped_name in ISP_REMAP_RULES:
            if keyword in normalized_input:
                jp_isp = mapped_name
                found = True
                break
        
        # 3. どのルールにも合致しなければ、元の英語名を返す

    jp_country = COUNTRY_JP_NAME.get(country_code, country_code)
    return jp_isp, jp_country

@st.cache_resource
def get_session():
    session = requests.Session()
    session.headers.update({"User-Agent": "WhoisBatchTool/2.1 (+PythonStreamlitApp)"})
    return session

session = get_session()

@st.cache_data
def get_world_map_data():
    try:
        world_geojson = alt.topo_feature('https://cdn.jsdelivr.net/npm/vega-datasets@v1.29.0/data/world-110m.json', 'countries')
        return world_geojson
    except Exception as e:
        st.error(f"GeoJSONデータのロード中にエラーが発生しました: {e}")
        return None

WORLD_MAP_GEOJSON = get_world_map_data()


# --- ヘルパー関数群 ---
def clean_ocr_error_chars(target):
    cleaned_target = target
    cleaned_target = cleaned_target.replace('Ⅱ', '11')
    cleaned_target = cleaned_target.replace('I', '1')
    cleaned_target = cleaned_target.replace('l', '1')
    cleaned_target = cleaned_target.replace('|', '1')
    cleaned_target = cleaned_target.replace('O', '0')
    cleaned_target = cleaned_target.replace('o', '0')
    cleaned_target = cleaned_target.replace('S', '5')
    cleaned_target = cleaned_target.replace('s', '5')
    if ':' not in cleaned_target:
        cleaned_target = cleaned_target.replace('A', '4')
        cleaned_target = cleaned_target.replace('a', '4')
        cleaned_target = cleaned_target.replace('B', '8')
    return cleaned_target

def is_valid_ip(target):
    try:
        ipaddress.ip_address(target)
        return True
    except ValueError:
        return False

def is_ipv4(target):
    try:
        ipaddress.IPv4Address(target)
        return True
    except ValueError:
        return False

def ip_to_int(ip):
    try:
        if is_ipv4(ip):
            return struct.unpack("!I", socket.inet_aton(ip))[0]
        return 0
    except OSError:
        return 0

def get_cidr_block(ip, netmask_range=(8, 24)):
    try:
        ip_obj = ipaddress.ip_address(ip)
        if ip_obj.version == 4:
            netmask = netmask_range[1] 
            network = ipaddress.ip_network(f'{ip}/{netmask}', strict=False)
            return str(network)
        elif ip_obj.version == 6:
            netmask = 48
            network = ipaddress.ip_network(f'{ip}/{netmask}', strict=False)
            return str(network)
        return None
    except ValueError:
        return None

def get_authoritative_rir_link(ip, country_code):
    rir_name = COUNTRY_CODE_TO_RIR.get(country_code)
    if rir_name and rir_name in RIR_LINKS:
        encoded_ip = quote(ip, safe='')
        if rir_name in ['RIPE', 'ARIN']:
            link_url = RIR_LINKS[rir_name].format(ip=encoded_ip)
            return f"[{rir_name}]({link_url})"
        elif rir_name in ['JPNIC', 'APNIC', 'LACNIC', 'AFRINIC']:
            link_url = RIR_LINKS[rir_name]  
            return f"[{rir_name} (手動検索)]({link_url})"
    return f"[Whois (汎用検索 - APNIC窓口)]({RIR_LINKS.get('APNIC', 'https://wq.apnic.net/static/search.html')})"

def get_copy_target(ip_display):
    if not ip_display: return ""
    return str(ip_display).split(' - ')[0].split(' ')[0]

def create_secondary_links(target):
    encoded_target = quote(target, safe='')
    is_ip = is_valid_ip(target)
    is_ipv6 = is_ip and not is_ipv4(target)

    who_is_url = f'https://who.is/whois-ip/ip-address/{encoded_target}' if is_ip else f'https://who.is/whois/{encoded_target}'
    dns_checker_url = ''
    dns_checker_key = ''

    if is_ip:
        dns_checker_path = 'ipv6-whois-lookup.php' if is_ipv6 else 'ip-whois-lookup.php'
        dns_checker_url = f'https://dnschecker.org/{dns_checker_path}?query={encoded_target}'
        dns_checker_key = 'DNS Checker (手動 - IPv6)' if is_ipv6 else 'DNS Checker'
    else:
        dns_checker_url = f'https://dnschecker.org/whois-lookup.php?query={encoded_target}'
        dns_checker_key = 'DNS Checker (ドメイン)'

    all_links = {
        'VirusTotal': f'https://www.virustotal.com/gui/search/{encoded_target}',
        'Aguse': f'https://www.aguse.jp/?url={encoded_target}',
        'Whois.com': f'https://www.whois.com/whois/{encoded_target}',
        'DomainSearch.jp': f'https://www.domainsearch.jp/whois/?q={encoded_target}',
        'Who.is': who_is_url,
        'IP2Proxy': f'https://www.ip2proxy.com/{encoded_target}',
        'DNSlytics (手動)': 'https://dnslytics.com/whois-lookup/',
        'IP Location': f'https://iplocation.io/ip/{encoded_target}',
        'CP-WHOIS (手動)': 'https://doco.cph.jp/whoisweb.php',
    }

    if dns_checker_url:
        all_links[dns_checker_key] = dns_checker_url

    if is_ipv6:
        links = {
            'VirusTotal': all_links['VirusTotal'],
            'DomainSearch.jp': all_links['DomainSearch.jp'],
            dns_checker_key: all_links[dns_checker_key],
            'IP2Proxy': all_links['IP2Proxy'],
            'DNSlytics (手動)': all_links['DNSlytics (手動)'],
            'IP Location': all_links['IP Location'],
            'CP-WHOIS (手動)': all_links['CP-WHOIS (手動)'],
        }
    else:
        links = all_links

    link_html = ""
    for name, url in links.items():
        link_html += f"[{name}]({url}) | "
    return link_html.rstrip(' | ')

# 🆕 RDAPデータ取得関数 (公式台帳への照会)
def fetch_rdap_data(ip):
    try:
        url = RDAP_BOOTSTRAP_URL.format(ip=ip)
        # RDAPはリダイレクトされることが多いため allow_redirects=True
        response = session.get(url, timeout=5, allow_redirects=True)
        if response.status_code == 200:
            data = response.json()
            # RDAPのレスポンス形式から組織名を探す (nameやremarks)
            network_name = data.get('name', '')
            
            # 詳細な記述がある場合もあるのでremarksも見るが、まずはnameを採用
            return network_name
    except:
        pass
    return None

# 🆕 Proモード用 API取得関数 (ipinfo.io) - 改良版
def get_ip_details_pro(ip, token, tor_nodes):
    result = {
        'Target_IP': ip, 'ISP': 'N/A', 'ISP_JP': 'N/A', 'Country': 'N/A', 'Country_JP': 'N/A', 
        'CountryCode': 'N/A', 'RIR_Link': 'N/A', 'Secondary_Security_Links': 'N/A', 'Status': 'N/A'
    }
    try:
        url = IPINFO_API_URL.format(ip=ip, token=token)
        response = session.get(url, timeout=10)
        
        if response.status_code == 429:
             result['Status'] = 'Error: Rate Limit (Pro)'
             return result

        response.raise_for_status()
        data = response.json()
        
        # --- 基本情報の取得 ---
        org_raw = data.get('org', '')
        isp_name = re.sub(r'^AS\d+\s+', '', org_raw) if org_raw else 'N/A'
        
        result['ISP'] = isp_name
        result['CountryCode'] = data.get('country', 'N/A')
        result['Country'] = result['CountryCode']
        result['RIR_Link'] = get_authoritative_rir_link(ip, result['CountryCode'])
        result['Status'] = 'Success (Pro API)'
        
        # 名寄せ処理
        jp_isp, jp_country = get_jp_names(result['ISP'], result['CountryCode'])
        result['ISP_JP'] = jp_isp
        result['Country_JP'] = jp_country

        # --- 🛡️ 判定ロジックの分岐 (ipinfoのprivacy情報を活用) ---
        privacy_data = data.get('privacy', {})
        
        if privacy_data:
            # --- パターンA: ipinfoの公式判定を採用 ---
            detected_types = []
            if privacy_data.get('vpn', False):
                detected_types.append("VPN")
            if privacy_data.get('proxy', False):
                detected_types.append("Proxy")
            if privacy_data.get('tor', False):
                detected_types.append("Tor Node")
            if privacy_data.get('hosting', False):
                detected_types.append("Hosting")
            if privacy_data.get('relay', False): # iCloud Private Relayなど
                detected_types.append("Relay")
            
            if detected_types:
                # 複数の性質を持つ場合もあるので結合 (例: "VPN, Hosting")
                result['Proxy_Type'] = ", ".join(detected_types)
            else:
                result['Proxy_Type'] = "" # 何も検知されなければ一般回線扱い
                
        else:
            # --- パターンB: privacyデータがない場合 (無料プラン等) ---
            # 従来通り、ツール独自のISP名判定ロジックを使用
            proxy_type = detect_proxy_vpn_tor(ip, result['ISP'], tor_nodes)
            is_anonymous = (proxy_type != "Standard Connection")
            result['Proxy_Type'] = f"{proxy_type}" if is_anonymous else ""
        
    except Exception as e:
        result['Status'] = f'Error: Pro API ({type(e).__name__})'
    
    result['Secondary_Security_Links'] = create_secondary_links(ip)
    return result

# --- API通信関数 (Main) ---
def get_ip_details_from_api(ip, cidr_cache_snapshot, delay_between_requests, rate_limit_wait_seconds, tor_nodes, use_rdap, api_key=None):
    
    # 1. Proモード (APIキーあり) の場合
    if api_key:
        result = get_ip_details_pro(ip, api_key, tor_nodes)
        
        # RDAPオプションが有効なら、Proモードでも追記する
        if use_rdap:
            rdap_result = fetch_rdap_data(ip)
            if rdap_result:
                result['ISP'] = f"{result['ISP']} [RDAP: {rdap_result}]"
        
        return result, None

    # 2. 通常モード (ip-api.com)
    result = {
        'Target_IP': ip, 'ISP': 'N/A', 'ISP_JP': 'N/A', 'Country': 'N/A', 'Country_JP': 'N/A', 
        'CountryCode': 'N/A', 'RIR_Link': 'N/A', 'Secondary_Security_Links': 'N/A', 'Status': 'N/A'
    }
    new_cache_entry = None

    cidr_block = get_cidr_block(ip)
    
    if cidr_block and cidr_block in cidr_cache_snapshot:
        cached_data = cidr_cache_snapshot[cidr_block]
        if time.time() - cached_data['Timestamp'] < 86400:
            result['ISP'] = cached_data['ISP']
            result['Country'] = cached_data['Country']
            result['CountryCode'] = cached_data['CountryCode']
            result['Status'] = "Success (Cache)" 
            jp_isp, jp_country = get_jp_names(result['ISP'], result['CountryCode'])
            proxy_type = detect_proxy_vpn_tor(ip, result['ISP'], tor_nodes)
            is_anonymous = (proxy_type != "Standard Connection")
            result['ISP_JP'] = jp_isp
            result['Proxy_Type'] = f"{proxy_type}" if is_anonymous else ""
            result['Country_JP'] = jp_country
            return result, None 

    try:
        time.sleep(delay_between_requests) 

        url = IP_API_URL.format(ip=ip)
        response = session.get(url, timeout=45)
        
        if response.status_code == 429:
            defer_until = time.time() + rate_limit_wait_seconds
            result['Status'] = 'Error: Rate Limit (429)'
            result['Defer_Until'] = defer_until
            result['Secondary_Security_Links'] = create_secondary_links(ip)
            return result, new_cache_entry 
        
        response.raise_for_status()
        data = response.json()
        
        if data.get('status') == 'success':
            country_code = data.get('countryCode', 'N/A') 

            # ISP名またはOrg名を採用
            raw_isp = data.get('isp', 'N/A')
            raw_org = data.get('org', '')
            
            # Org情報がある場合は、ISP情報と併記または優先度判定
            combined_name = raw_isp
            if raw_org and raw_org != raw_isp:
                combined_name = f"{raw_isp} / {raw_org}"
            
            result['ISP'] = combined_name
            result['Country'] = data.get('country', 'N/A')
            result['CountryCode'] = data.get('countryCode', 'N/A')
            result['RIR_Link'] = get_authoritative_rir_link(ip, country_code)
            status_type = "IPv6 API" if not is_ipv4(ip) else "IPv4 API"
            
            # --- 🆕 RDAP併用ロジック ---
            rdap_name = ""
            if use_rdap:
                rdap_result = fetch_rdap_data(ip)
                if rdap_result:
                    rdap_name = rdap_result
                    result['ISP'] = f"{result['ISP']} [RDAP: {rdap_name}]"

            result['Status'] = f'Success ({status_type})'
            
            # 名寄せ処理
            jp_isp, jp_country = get_jp_names(result['ISP'], country_code)
            
            proxy_type = detect_proxy_vpn_tor(ip, result['ISP'], tor_nodes)
            is_anonymous = (proxy_type != "Standard Connection")
            result['ISP_JP'] = jp_isp
            result['Proxy_Type'] = f"{proxy_type}" if is_anonymous else ""
            result['Country_JP'] = jp_country
            
            if cidr_block:
                new_cache_entry = {
                    cidr_block: {
                    'ISP': result['ISP'], # RDAP情報込みでキャッシュする
                    'Country': result['Country'],
                    'CountryCode': result['CountryCode'],
                    'Timestamp': time.time()
                    }
                }
            
        elif data.get('status') == 'fail':
            result['Status'] = f"API Fail: {data.get('message', 'Unknown Fail')}"
            result['RIR_Link'] = get_authoritative_rir_link(ip, 'N/A')
            
        else:
            result['Status'] = f"API Error: Unknown Response"
            result['RIR_Link'] = get_authoritative_rir_link(ip, 'N/A')
            
    except requests.exceptions.RequestException as e:
        result['Status'] = f'Error: Network/Timeout ({type(e).__name__})'
        
    result['Secondary_Security_Links'] = create_secondary_links(ip)
    return result, new_cache_entry

def get_domain_details(domain):
    icann_link = f"[ICANN Whois (手動検索)]({RIR_LINKS['ICANN Whois']})"
    return {
        'Target_IP': domain, 'ISP': 'Domain/Host', 'Country': 'N/A', 'CountryCode': 'N/A',
        'RIR_Link': icann_link,
        'Secondary_Security_Links': create_secondary_links(domain),
        'Status': 'Success (Domain)'
    }

def get_simple_mode_details(target):
    if is_valid_ip(target):
        rir_link_content = f"[Whois (汎用検索 - APNIC窓口)]({RIR_LINKS['APNIC']})"
    else:
        rir_link_content = f"[ICANN Whois (手動検索)]({RIR_LINKS['ICANN Whois']})"
        
    return {
        'Target_IP': target, 
        'ISP': 'N/A (簡易モード)', 
        'Country': 'N/A (簡易モード)',
        'CountryCode': 'N/A',
        'RIR_Link': rir_link_content,
        'Secondary_Security_Links': create_secondary_links(target),
        'Status': 'Success (簡易モード)' 
    }

# --- ヘルパー関数群 ---

def group_results_by_isp(results):
    grouped = {}
    final_grouped_results = []
    non_aggregated_results = []
    successful_results = [res for res in results if res['Status'].startswith('Success')]

    for res in successful_results:
        is_ip = is_valid_ip(res['Target_IP'])
        if not is_ip or not is_ipv4(res['Target_IP']) or res['ISP'] == 'N/A' or res['Country'] == 'N/A' or res['ISP'] == 'N/A (簡易モード)':
            if res['Status'].startswith('Success (IPv4 CIDR Cache)'):
                non_aggregated_results.append(res)
            else:
                non_aggregated_results.append(res)
            continue
        
        key = (res['ISP'], res['CountryCode']) 
        
        if key not in grouped:
            grouped[key] = {
                'IP_Ints': [], 'IPs_List': [], 'RIR_Link': res['RIR_Link'],
                'Secondary_Security_Links': res['Secondary_Security_Links'],
                'ISP': res['ISP'], 
                'Country': res['Country'], 
                'Status': res['Status'],
                'ISP_JP': res.get('ISP_JP', 'N/A'),
                'Country_JP': res.get('Country_JP', 'N/A')
            }
        ip_int = ip_to_int(res['Target_IP'])
        if ip_int != 0:
            grouped[key]['IP_Ints'].append(ip_int)
            grouped[key]['IPs_List'].append(res['Target_IP'])
        else:
            res['Status'] = 'Error: IPv4 Int Conversion Failed'
            non_aggregated_results.append(res)

    non_aggregated_results.extend([res for res in results if not res['Status'].startswith('Success')])
    
    for key, data in grouped.items():
        if not data['IP_Ints']: 
            continue
            
        sorted_ip_ints = sorted(data['IP_Ints'])
        min_int = sorted_ip_ints[0]
        max_int = sorted_ip_ints[-1]
        count = len(data['IPs_List'])
        try:
            min_ip = str(ipaddress.IPv4Address(min_int))
            max_ip = str(ipaddress.IPv4Address(max_int))
        except ValueError:
            min_ip = data['IPs_List'][0]
            max_ip = data['IPs_List'][-1]
        
        target_ip_display = min_ip if count == 1 else f"{min_ip} - {max_ip} (x{count} IPs)"
        status_display = data['Status'] if count == 1 else f"Aggregated ({count} IPs)"
        
        final_grouped_results.append({
            'Target_IP': target_ip_display, 
            'Country': data['Country'], 
            'Country_JP': data['Country_JP'], 
            'ISP': data['ISP'],
            'ISP_JP': data['ISP_JP'], 
            'RIR_Link': data['RIR_Link'], 
            'Secondary_Security_Links': data['Secondary_Security_Links'],
            'Status': status_display
        })
    
    final_grouped_results.extend(non_aggregated_results)

    return final_grouped_results

# --- リアルタイム集計関数 ---
def summarize_in_realtime(raw_results):
    isp_counts = {}
    country_counts = {}
    country_code_counts = {}

    target_frequency = st.session_state.get('target_freq_map', {})

    st.session_state['debug_summary'] = {} 

    country_all_df_raw = pd.DataFrame({
        'NumericCode': pd.Series(dtype='int64'), 
        'Count': pd.Series(dtype='int64'),
        'Country': pd.Series(dtype='str')
    })

    success_ipv4 = [
        r for r in raw_results 
        if r['Status'].startswith('Success') and is_ipv4(r['Target_IP'])
    ]

    for r in success_ipv4:
        ip = r.get('Target_IP')
        frequency = target_frequency.get(ip, 1) 

        isp_name = r.get('ISP_JP', r.get('ISP', 'N/A'))
        country_name = r.get('Country_JP', r.get('Country', 'N/A'))
        cc = r.get('CountryCode', 'N/A')
        
        if isp_name and isp_name not in ['N/A', 'N/A (簡易モード)']:
            isp_counts[isp_name] = isp_counts.get(isp_name, 0) + frequency
        
        if country_name and country_name != 'N/A':
            country_counts[country_name] = country_counts.get(country_name, 0) + frequency
            
        if cc and cc != 'N/A':
            country_code_counts[cc] = country_code_counts.get(cc, 0) + frequency

    # --- ISP集計 ---
    isp_full_df = pd.DataFrame(list(isp_counts.items()), columns=['ISP', 'Count'])
    isp_full_df = isp_full_df.sort_values('Count', ascending=False)
    
    if not isp_full_df.empty:
        isp_df = isp_full_df.head(10).copy()
        isp_df['ISP'] = isp_df['ISP'].str.wrap(25)
    else:
        isp_df = pd.DataFrame(columns=['ISP', 'Count'])

    # --- 国集計 ---
    country_full_df = pd.DataFrame(list(country_counts.items()), columns=['Country', 'Count'])
    country_full_df = country_full_df.sort_values('Count', ascending=False)

    if not country_full_df.empty:
        country_df = country_full_df.head(10).copy()
        country_df['Country'] = country_df['Country'].str.wrap(25)
    else:
        country_df = pd.DataFrame(columns=['Country', 'Count'])

    # ヒートマップ用
    if country_code_counts:
        map_data = []
        for cc, cnt in country_code_counts.items():
            num = COUNTRY_CODE_TO_NUMERIC_ISO.get(cc)
            if num is not None:
                name_for_map = COUNTRY_JP_NAME.get(cc, cc)
                map_data.append({
                    'NumericCode': int(num), 
                    'Count': int(cnt),
                    'Country': name_for_map
                })

        country_all_df_raw = pd.DataFrame(map_data).astype({
            'NumericCode': 'int64',
            'Count': 'int64'
        })
        
    st.session_state['debug_summary']['country_code_counts'] = country_code_counts
    st.session_state['debug_summary']['country_all_df'] = country_all_df_raw.to_dict('records')

    # --- ターゲット頻度集計 ---
    freq_map = st.session_state.get('target_freq_map', {})
    finished = st.session_state.get('finished_ips', set())
    freq_list = [{'Target_IP': t, 'Count': c} for t, c in freq_map.items() if t in finished]
    
    if freq_list:
        freq_full_df = pd.DataFrame(freq_list).sort_values('Count', ascending=False)
    else:
        freq_full_df = pd.DataFrame(columns=['Target_IP', 'Count'])
    
    if not freq_full_df.empty:
        freq_df = freq_full_df.head(10).copy()
    else:
        freq_df = pd.DataFrame(columns=['Target_IP', 'Count'])

    return isp_df, country_df, freq_df, country_all_df_raw, isp_full_df, country_full_df, freq_full_df

# --- 集計結果描画ヘルパー関数 ---
def draw_summary_content(isp_summary_df, country_summary_df, target_frequency_df, country_all_df, title):
    st.subheader(title)
    
    st.markdown("#### 🌍 国別 IP カウントヒートマップ")
    if WORLD_MAP_GEOJSON and not country_all_df.empty:
        
        base = alt.Chart(WORLD_MAP_GEOJSON).mark_geoshape(
            stroke='black', 
            strokeWidth=0.1
        ).encode(
            color=alt.value("#f0f0f052"), 
        ).project(
            type='mercator',
            scale=80,
            translate=[500, 180]        
        ).properties(
            title='World Map IP Count Heatmap',
            width=2500, 
            height=400 
        )

        heatmap = alt.Chart(WORLD_MAP_GEOJSON).mark_geoshape(
            stroke='black', 
            strokeWidth=0.1
        ).encode(
            color=alt.Color('Count:Q',
                            scale=alt.Scale(
                                type='log', 
                                domainMin=1,
                                domainMax=alt.Undefined,
                                range=['#99f6e4', '#facc15', '#dc2626']
                            ),
                            legend=alt.Legend(title="IP Count")),
            tooltip=[
                alt.Tooltip('Country:N', title='Country'),
                alt.Tooltip('Count:Q', title='IP Count', format=',')
            ]
        ).transform_lookup(
            lookup='id',
            from_=alt.LookupData(
                country_all_df, 
                key='NumericCode',          
                fields=['Count', 'Country']
            )
        ).project(
            type='mercator',
            scale=80,
            translate=[500, 180]
            )

        chart = alt.layer(base, heatmap).resolve_scale(
            color='independent'
        ).configure_legend(
            orient='right'
        ).interactive()
        
        st.altair_chart(chart, use_container_width=True)
        
    else:
        st.info("ヒートマップデータまたはGeoJSONがロードされていないか、成功したIPv4データが存在しないため表示できません。")
    
    st.markdown("---")


    col_freq, col_isp, col_country = st.columns([1, 1, 1]) 

    # 共通チャート生成関数
    def create_labeled_bar_chart(df, x_field, y_field, title):
        base = alt.Chart(df).encode(
            x=alt.X(x_field, title='Count'),
            y=alt.Y(y_field, sort='-x', title=y_field),
            tooltip=[y_field, x_field]
        )
        bars = base.mark_bar()
        text = base.mark_text(
            align='left',
            baseline='middle',
            dx=3 
        ).encode(
            text=x_field
        )
        return (bars + text).properties(title=title).interactive()

    with col_freq:
        st.markdown("#### 🎯 対象IP別カウント (トップ10)")
        if not target_frequency_df.empty:
            st.caption(f"**集計対象ターゲット数 (重複なし):** {len(target_frequency_df)} 件")
            chart = create_labeled_bar_chart(target_frequency_df, 'Count', 'Target_IP', 'Target IP Counts')
            st.altair_chart(chart, use_container_width=True)

            target_frequency_df_display = target_frequency_df.copy()
            target_frequency_df_display['Target_IP'] = target_frequency_df_display['Target_IP'].str.wrap(25)
            st.dataframe(target_frequency_df_display, hide_index=True, use_container_width=True)
        else:
            st.info("データがありません")
            
    with col_isp:
        st.markdown("#### 🏢 ISP別カウント (トップ10)")
        if not isp_summary_df.empty:
            chart = create_labeled_bar_chart(isp_summary_df, 'Count', 'ISP', 'ISP Counts')
            st.altair_chart(chart, use_container_width=True)
            
            st.dataframe(isp_summary_df, hide_index=True, use_container_width=True)
        else:
            st.info("データがありません")
            
    with col_country:
        st.markdown("#### 🌍 国別カウント (トップ10)")
        if not country_summary_df.empty:
            chart = create_labeled_bar_chart(country_summary_df, 'Count', 'Country', 'Country Counts')
            st.altair_chart(chart, use_container_width=True)
            
            st.dataframe(country_summary_df, hide_index=True, use_container_width=True)
        else:
            st.info("データがありません")

# 💡 HTMLレポート生成関数（改良版 v7.2）
def generate_full_report_html(isp_full_df, country_full_df, freq_full_df):
    
    def create_chunked_chart_specs(df, x_col, y_col, title_base, chunk_size=50):
        specs = []
        # データ全体での最大値を取得 (ページまたぎのスケール統一のため)
        global_max = df[x_col].max() if not df.empty else 0

        # データフレームを分割
        chunks = [df[i:i + chunk_size] for i in range(0, df.shape[0], chunk_size)]
        
        for i, chunk in enumerate(chunks):
            chart_title = f"{title_base} ({i+1}/{len(chunks)})" if len(chunks) > 1 else title_base
            
            # 数値ラベル付きチャート
            # 💡 x軸のスケールを全体最大値で固定する
            base = alt.Chart(chunk).encode(
                x=alt.X(x_col, title='Count', scale=alt.Scale(domain=[0, global_max])),
                y=alt.Y(y_col, sort='-x', title=y_col),
                tooltip=[y_col, x_col]
            )
            bars = base.mark_bar()
            text = base.mark_text(
                align='left',
                baseline='middle',
                dx=5, 
                fontSize=11,
                fontWeight='bold'
            ).encode(
                text=x_col
            )
            chart = (bars + text).properties(
                title=chart_title,
                width=700,
                height=alt.Step(20) 
            )
            specs.append(chart.to_dict())
        return specs

    # 各カテゴリのチャートスペックを生成
    target_specs = create_chunked_chart_specs(freq_full_df, 'Count', 'Target_IP', 'Target IP Counts (All)')
    isp_specs = create_chunked_chart_specs(isp_full_df, 'Count', 'ISP', 'ISP Counts (All)')
    country_specs = create_chunked_chart_specs(country_full_df, 'Count', 'Country', 'Country Counts (All)')

    # HTMLテンプレート
    html_template = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <title>Whois Search Full Report</title>
      <script src="https://cdn.jsdelivr.net/npm/vega@5"></script>
      <script src="https://cdn.jsdelivr.net/npm/vega-lite@5"></script>
      <script src="https://cdn.jsdelivr.net/npm/vega-embed@6"></script>
      <style>
        body {{ font-family: "Helvetica Neue", Arial, sans-serif; padding: 40px; background-color: #fff; color: #333; }}
        h1 {{ text-align: center; border-bottom: 2px solid #333; padding-bottom: 10px; margin-bottom: 30px; }}
        h2 {{ 
            color: #1e3a8a; 
            margin-top: 50px; 
            border-left: 5px solid #1e3a8a; 
            padding-left: 15px; 
            page-break-before: always; 
        }}
        h2:first-of-type {{ page-break-before: auto; }} 
        
        .chart-container {{ 
            margin-bottom: 40px; 
            padding: 10px; 
            page-break-inside: avoid; 
        }}
        
        @media print {{
            body {{ padding: 0; background-color: #fff; }}
            .no-print {{ display: none; }}
            h2 {{ margin-top: 20px; }}
        }}
      </style>
    </head>
    <body>
      <h1>Whois検索結果分析レポート</h1>
      <p style="text-align: center; color: #666;">Generated by Whois Search Tool</p>

      <h2>対象IPアドレス カウント (全 {len(freq_full_df)} 件)</h2>
      <div id="target_charts"></div>

      <h2>ISP別 カウント (全 {len(isp_full_df)} 件)</h2>
      <div id="isp_charts"></div>

      <h2>国別 カウント (全 {len(country_full_df)} 件)</h2>
      <div id="country_charts"></div>

      <script type="text/javascript">
        // Embed charts function
        function embedCharts(containerId, specs) {{
            const container = document.getElementById(containerId);
            specs.forEach((spec, index) => {{
                const div = document.createElement('div');
                div.id = containerId + '_' + index;
                div.className = 'chart-container';
                container.appendChild(div);
                vegaEmbed('#' + div.id, spec, {{actions: false}});
            }});
        }}

        // Data from Python (Serialized to JSON)
        const targetSpecs = {json.dumps(target_specs)};
        const ispSpecs = {json.dumps(isp_specs)};
        const countrySpecs = {json.dumps(country_specs)};

        // Render
        if (targetSpecs.length > 0) embedCharts('target_charts', targetSpecs);
        else document.getElementById('target_charts').innerHTML = '<p>データなし</p>';

        if (ispSpecs.length > 0) embedCharts('isp_charts', ispSpecs);
        else document.getElementById('isp_charts').innerHTML = '<p>データなし</p>';

        if (countrySpecs.length > 0) embedCharts('country_charts', countrySpecs);
        else document.getElementById('country_charts').innerHTML = '<p>データなし</p>';
      </script>
    </body>
    </html>
    """
    return html_template

# 📈 クロス分析用HTMLレポート生成関数
def generate_cross_analysis_html(chart_spec, x_col, group_col):
    html_template = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <title>Whois Cross Analysis Report</title>
      <script src="https://cdn.jsdelivr.net/npm/vega@5"></script>
      <script src="https://cdn.jsdelivr.net/npm/vega-lite@5"></script>
      <script src="https://cdn.jsdelivr.net/npm/vega-embed@6"></script>
      <style>
        body {{ font-family: "Helvetica Neue", Arial, sans-serif; padding: 40px; background-color: #fff; color: #333; }}
        h1 {{ text-align: center; border-bottom: 2px solid #333; padding-bottom: 10px; margin-bottom: 30px; }}
        .info {{ text-align: center; color: #666; margin-bottom: 20px; }}
        .chart-container {{ 
            width: 100%; 
            display: flex; 
            justify-content: center; 
            margin-bottom: 40px; 
            padding: 10px; 
        }}
      </style>
    </head>
    <body>
      <h1>クロス分析レポート: {x_col} vs {group_col}</h1>
      <p class="info">Generated by Whois Search Tool</p>

      <div id="chart" class="chart-container"></div>

      <script type="text/javascript">
        const spec = {json.dumps(chart_spec)};
        vegaEmbed('#chart', spec, {{actions: true}});
      </script>
    </body>
    </html>
    """
    return html_template

# --- Excel生成ヘルパー関数 ---
def convert_df_to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    return output.getvalue()

# --- Advanced Excel Generator (Pivot & Chart) v5.0 ---
def create_advanced_excel(df, time_col_name=None):
    """
    1. Raw Data
    2. Report_ISP_Volume: ISP Ranking (Bar Chart)
    3. Report_ISP_Risk: ISP x ProxyType (Stacked Bar)
    4. Report_Time_Volume: Hour Trend (Bar/Line) [if time_col available]
    5. Report_Time_Risk: Hour x ProxyType (Stacked Bar) [if time_col available]
    """
    output = io.BytesIO()
    
    # 1. データ前処理
    # Proxy Typeの空欄を「Standard Connection」で埋める (用語変更)
    if 'Proxy Type' in df.columns:
        df['Proxy Type'] = df['Proxy Type'].fillna('Standard Connection')
        df['Proxy Type'] = df['Proxy Type'].replace('', 'Standard Connection')
        # 古い用語が残っている場合の念の為の置換
        df['Proxy Type'] = df['Proxy Type'].replace('Residential/Normal', 'Standard Connection')
        df['Proxy Type'] = df['Proxy Type'].replace('Residential/General', 'Standard Connection')
        df['Proxy Type'] = df['Proxy Type'].replace('Residential/Business', 'Standard Connection')
        df['Proxy Type'] = df['Proxy Type'].replace('nan', 'Standard Connection')
    else:
        df['Proxy Type'] = 'Standard Connection'
    
    # 時間帯列の作成
    has_time_analysis = False
    if time_col_name and time_col_name in df.columns:
        try:
            df['Hour'] = pd.to_datetime(df[time_col_name], errors='coerce').dt.hour
            has_time_analysis = True
        except Exception:
            pass

    # カウント用の列（最初の列を使う）
    count_col = df.columns[0]

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Raw Data
        df.to_excel(writer, index=False, sheet_name='Raw Data')
        wb = writer.book
        
        # --- 共通チャート作成関数 (解説文付き) ---
        def add_chart_sheet(pivot_df, sheet_name, chart_title, x_title, y_title, description, chart_type="col", stacked=False):
            if pivot_df.empty: return

            # Sheet作成とデータ書き込み (ヘッダー用に少し下げる)
            pivot_df.to_excel(writer, sheet_name=sheet_name, startrow=4)
            ws = wb[sheet_name]
            
            # --- 解説文の挿入 ---
            ws['A1'] = chart_title
            ws['A1'].font = Font(size=14, bold=True, color="1E3A8A")
            
            ws['A2'] = description
            ws['A2'].font = Font(size=11, color="555555", italic=True)
            ws['A2'].alignment = Alignment(wrap_text=True, vertical="top")
            
            # セル結合 (説明文エリア)
            ws.merge_cells('A2:H3')
            
            # 印刷設定（横向き）
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.print_options.horizontalCentered = True
            
            # グラフ作成
            chart = BarChart()
            chart.type = chart_type
            chart.style = 10
            chart.title = chart_title
            chart.y_axis.title = y_title
            chart.x_axis.title = x_title
            if stacked:
                chart.grouping = "stacked"
                chart.overlap = 100
            
            chart.height = 15 # 印刷用に見やすく大きく
            chart.width = 25

            # データ範囲設定 (startrow=4 なのでデータは5行目から)
            # ヘッダーは 5行目
            # データ開始は 6行目
            data_start_row = 5 
            data_end_row = data_start_row + len(pivot_df)
            
            data = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row, max_col=len(pivot_df.columns)+1)
            cats = Reference(ws, min_col=1, min_row=data_start_row+1, max_row=data_end_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # グラフ配置 (データの下ではなく横に配置して見やすく)
            ws.add_chart(chart, "E5")

        # ---------------------------------------------------------
        # 2. Report_ISP_Volume: [ISP_JP] x [Count]
        # ---------------------------------------------------------
        top_isps = df['ISP_JP'].value_counts().head(20).index
        df_isp = df[df['ISP_JP'].isin(top_isps)]
        pivot_isp_vol = df_isp.pivot_table(
            index='ISP_JP', 
            values=count_col, 
            aggfunc='count'
        ).sort_values(count_col, ascending=False)
        
        desc_isp_vol = "どのプロバイダからのアクセスが最も多いかを可視化しています。特定のISPからのアクセス集中は、そのサービスの利用者層または特定のキャンペーンの影響を示唆します。"
        add_chart_sheet(pivot_isp_vol, 'Report_ISP_Volume', 'ISP Access Volume Ranking (Top 20)', 'Internet Service Provider', 'Access Count (件数)', desc_isp_vol)

        # ---------------------------------------------------------
        # 3. Report_ISP_Risk: [ISP_JP] x [Proxy Type]
        # ---------------------------------------------------------
        pivot_isp_risk = df_isp.pivot_table(
            index='ISP_JP', 
            columns='Proxy Type', 
            values=count_col, 
            aggfunc='count', 
            fill_value=0
        )
        desc_isp_risk = "そのISPが安全な一般回線か、注意が必要なサーバー/VPN経由かを判定しています。「Standard Connection」は一般的な安全な接続です。「Hosting」や「VPN」が多い場合は機械的なアクセスの可能性があります。"
        add_chart_sheet(pivot_isp_risk, 'Report_ISP_Risk', 'Risk Analysis by ISP (Top 20)', 'Internet Service Provider', 'Access Count (件数)', desc_isp_risk, stacked=True)
        
        # ---------------------------------------------------------
        # 4. Report_Country: [Country_JP] x [Count] (Bonus)
        # ---------------------------------------------------------
        pivot_country = df.pivot_table(
            index='Country_JP',
            values=count_col,
            aggfunc='count'
        ).sort_values(count_col, ascending=False).head(15)
        desc_country = "国ごとのアクセス数をランキング化しています。サービス提供エリア外からの予期せぬアクセス検知や、海外からの攻撃予兆の発見に役立ちます。"
        add_chart_sheet(pivot_country, 'Report_Country', 'Country Access Volume (Top 15)', 'Country Name', 'Access Count (件数)', desc_country)

        # ---------------------------------------------------------
        # 5. Time Analysis (if available)
        # ---------------------------------------------------------
        if has_time_analysis:
            # Report_Time_Volume: [Hour] x [Count]
            pivot_time_vol = df.pivot_table(
                index='Hour',
                values=count_col,
                aggfunc='count',
                fill_value=0
            ).reindex(range(24), fill_value=0)
            desc_time_vol = "何時にアクセスが集中しているかを可視化しています。一般的なユーザーは活動時間帯に、Botなどは深夜早朝や24時間一定のアクセスを行う傾向があります。"
            add_chart_sheet(pivot_time_vol, 'Report_Time_Volume', 'Hourly Access Trend', 'Time of Day (0-23h)', 'Access Count (件数)', desc_time_vol)

            # Report_Time_Risk: [Hour] x [Proxy Type]
            pivot_time_risk = df.pivot_table(
                index='Hour',
                columns='Proxy Type',
                values=count_col,
                aggfunc='count',
                fill_value=0
            ).reindex(range(24), fill_value=0)
            desc_time_risk = "深夜帯などに怪しいアクセス（Hosting/VPN等）が増えていないかを確認できます。夜間にHosting判定が増加する場合、Botによる自動巡回の可能性があります。"
            add_chart_sheet(pivot_time_risk, 'Report_Time_Risk', 'Hourly Risk Trend', 'Time of Day (0-23h)', 'Access Count (件数)', desc_time_risk, stacked=True)
            
    return output.getvalue()


def display_results(results, current_mode_full_text, display_mode):
    st.markdown("### 📝 検索結果")

    with st.expander("⚠️ 判定アイコンと表示ルールについて"):
        st.info("""
        ### 🔍 判定ロジックの概要
        本ツールは、IPアドレスに紐付けられた**ASN（Autonomous System Number）およびISP（インターネットサービスプロバイダ）の名称・属性**を解析し、通信主体のネットワーク種別を自動的に分類しています。
        
        インターネット上の通信は、その用途に応じて「個人宅・法人拠点からの直接接続」と「非対面的な中継・ホスティング経由の接続」に大別されます。本機能は後者を検知し、調査の優先順位判断を支援することを目的としています。
        
        ---
        
        ### 📌 判定種別の定義と技術的背景
        
        - **⚠️ [Tor Node]**
            - **定義**: Tor（The Onion Router）ネットワークにおける「Exit Node（出口ノード）」を指します。
            - **背景**: 起動時にTor Project公式サイトより最新のノードリストを取得し、照合を行っています。高い匿名性を維持した通信であるため、セキュリティリスクの検討が必要です。
            
        - **⚠️ [VPN/Proxy]**
            - **定義**: 商用VPNサービス、公開プロキシ、またはプライバシー保護を目的とした中継団体に属するIPです。
            - **背景**: ISP名称に含まれる特定のキーワード（VPN, Proxy等）および既知の匿名化サービス運営組織名に基づき判別します。
            
        - **⚠️ [Hosting/Infra]**
            - **定義**: クラウドサービス（AWS, Azure, GCP等）や、データセンター、ホスティング事業者のインフラストラクチャです。
            - **背景**: 一般的なコンシューマ回線とは異なり、サーバー間通信やBot、クローラー、あるいは攻撃用インフラとして利用されるケースが多いノードです。
            
        ---
        
        ※ 本判定はISP名称等に基づく推論であるため、実際の利用状況と異なる場合があります。
        """)
    
    col_widths = [0.5, 1.5, 1.2, 2.0, 1.5, 1.5, 1.0, 1.2, 0.5] 
    h_cols = st.columns(col_widths)
    headers = ["No.", "Target IP", "国名","ISP(日本語)", "RIR Link", "Security Links", "Proxy Type",  "Status", "✅"]
    for col, name in zip(h_cols, headers):
        col.markdown(f"**{name}**")
    st.markdown("<hr style='margin: 0px 0px 10px 0px;'>", unsafe_allow_html=True)

    with st.container(height=800):
        if not results:
            st.info("検索結果がここに表示されます。")
            return

        for idx, res in enumerate(results):
                row_cols = st.columns(col_widths)
                row_cols[0].write(f"**{idx+1}**")
                
                target_ip = res.get('Target_IP', 'N/A')
                row_cols[1].markdown(f"`{target_ip}`")
                
                c_jp = res.get('Country_JP', 'N/A')
                c_en = res.get('Country', 'N/A')
                row_cols[2].write(f"{c_jp}\n({c_en})")
                
                isp_display = res.get('ISP_JP', res.get('ISP', 'N/A'))
                row_cols[3].write(isp_display)
                
                rir_link = res.get('RIR_Link', 'N/A')
                with row_cols[4]:
                    st.write(rir_link)
                    clean_ip = get_copy_target(target_ip)
                    st.code(clean_ip, language=None)
                
                row_cols[5].write(res.get('Secondary_Security_Links', 'N/A'))
                hosting_val = res.get('Proxy_Type', '')
                row_cols[6].write(hosting_val)          
                
                status_val = res.get('Status', 'N/A')
                if "Success" in status_val:
                    row_cols[7].markdown(f"<span style='color:green;'>{status_val}</span>", unsafe_allow_html=True)
                else:
                    row_cols[7].write(status_val)
                    
                row_cols[8].checkbox("選択", key=f"chk_{get_copy_target(target_ip)}_{idx}", label_visibility="collapsed")


# 📊 元データ結合分析機能
def render_merged_analysis(df_merged):
    st.markdown("### 📈 元データ x 検索結果 クロス分析")
    st.info("アップロードされたファイルの元の列と、検索で得られたWhois情報を組み合わせて可視化します。印刷用にグラフ単体のダウンロードも可能です。")
    
    # グラフ設定用カラム
    # 元データのカラム（Statusなど後付けのカラムを除く）
    original_cols = [c for c in df_merged.columns if c not in ['ISP', 'ISP_JP', 'Country', 'Country_JP', 'Proxy Type', 'Status']]
    # Whois結果のカラム
    whois_cols = ['Country_JP', 'ISP_JP', 'Proxy Type', 'Status']
    
    col_x, col_grp, col_chart_type = st.columns(3)
    
    with col_x:
        x_col = st.selectbox("X軸 (カテゴリ/元の列)", original_cols + whois_cols, index=0)
    
    with col_grp:
        group_col = st.selectbox("積み上げ/色分け (Whois情報など)", ['(なし)'] + whois_cols + original_cols, index=1)
        
    with col_chart_type:
        chart_type = st.radio("グラフタイプ", ["バーチャート (集計)", "ヒートマップ"], horizontal=True)

    if not df_merged.empty:
        chart = None
        
        # データ前処理: NaNを文字列に置換してAltairのエラー回避
        chart_df = df_merged.fillna("N/A").astype(str)

        if chart_type == "バーチャート (集計)":
            if group_col != '(なし)':
                chart = alt.Chart(chart_df).mark_bar().encode(
                    x=alt.X(x_col, title=x_col),
                    y=alt.Y('count()', title='件数'),
                    color=alt.Color(group_col, title=group_col),
                    tooltip=[x_col, group_col, 'count()']
                ).properties(height=400)
            else:
                chart = alt.Chart(chart_df).mark_bar().encode(
                    x=alt.X(x_col, title=x_col),
                    y=alt.Y('count()', title='件数'),
                    tooltip=[x_col, 'count()']
                ).properties(height=400)
                
        elif chart_type == "ヒートマップ":
             if group_col != '(なし)':
                chart = alt.Chart(chart_df).mark_rect().encode(
                    x=alt.X(x_col, title=x_col),
                    y=alt.Y(group_col, title=group_col),
                    color=alt.Color('count()', title='件数', scale=alt.Scale(scheme='viridis')),
                    tooltip=[x_col, group_col, 'count()']
                ).properties(height=400)
             else:
                 st.warning("ヒートマップには「積み上げ/色分け」項目の選択が必要です。")

        if chart:
            st.altair_chart(chart, use_container_width=True)
            
            # HTMLダウンロード用
            chart_json = chart.to_dict()
            html_content = generate_cross_analysis_html(chart_json, x_col, group_col if group_col != '(なし)' else 'Count')
            
            st.download_button(
                label="⬇️ クロス分析レポート(HTML)をダウンロード",
                data=html_content,
                file_name=f"cross_analysis_{x_col}_vs_{group_col}.html",
                mime="text/html",
                help="グラフをブラウザで全画面表示し、印刷するのに適しています。"
            )


# --- メイン処理 ---
def main():
    if 'cancel_search' not in st.session_state: st.session_state['cancel_search'] = False
    if 'raw_results' not in st.session_state: st.session_state['raw_results'] = []
    if 'targets_cache' not in st.session_state: st.session_state['targets_cache'] = []
    if 'is_searching' not in st.session_state: st.session_state['is_searching'] = False
    if 'deferred_ips' not in st.session_state: st.session_state['deferred_ips'] = {} 
    if 'finished_ips' not in st.session_state: st.session_state['finished_ips'] = set() 
    if 'search_start_time' not in st.session_state: st.session_state['search_start_time'] = 0.0 
    if 'target_freq_map' not in st.session_state: st.session_state['target_freq_map'] = {} 
    if 'cidr_cache' not in st.session_state: st.session_state['cidr_cache'] = {} 
    if 'debug_summary' not in st.session_state: st.session_state['debug_summary'] = {}

    tor_nodes = fetch_tor_exit_nodes()
    
    with st.sidebar:
        st.markdown("### 🛠️ Menu")
        selected_menu = option_menu(
            menu_title=None,
            options=["Whois検索", "仕様・解説"],
            icons=["search", "book"],
            default_index=0,
            styles={
                "nav-link": {"font-size": "16px", "text-align": "left", "margin": "5px", "--hover-color": "#eee"},
                "nav-link-selected": {"background-color": "#1e3a8a"},
            }
        )
        st.markdown("---")
        
        # 🆕 Proモード設定 (APIキー入力)
        st.markdown("#### 🔑 Pro Mode (Optional)")
        pro_api_key = st.text_input("ipinfo.io API Key", type="password", help="入力するとipinfo.ioの高精度データベースを使用します。空欄の場合はip-api.com(無料)を使用します。")
        
        st.markdown("---")
        if st.button("🔄 IPキャッシュクリア", help="キャッシュが古くなった場合にクリック"):
            st.session_state['cidr_cache'] = {} 
            st.info("IP/CIDRキャッシュをクリアしました。")
            st.rerun()

    if selected_menu == "仕様・解説":
        st.title("📖 マニュアル & ガイド")
        
        # タブで情報を整理して見やすくする
        tab1, tab2, tab3 = st.tabs(["🔰 使い方・モード選択", "⚙️ 仕様・技術詳細", "❓ FAQ"])
        
        with tab1:
            st.markdown("### 🚀 クイックスタート")
            st.markdown("""
            1. **入力**: 左側のテキストエリアにIPアドレスを貼り付けるか、テキストファイルをアップロードします。
            2. **設定**: 基本的にはそのままでOKです。より詳しい情報が必要な場合は下の表を参考にモードを変更してください。
            3. **実行**: 「🚀 検索開始」ボタンを押します。
            """)
            
            st.info("💡 **ヒント**: 結果が出たあと、画面下のボタンからExcelファイルをダウンロードすると、自動でグラフ化された分析レポートが見れます。")

            st.markdown("---")
            st.markdown("### 📊 目的別：モードの選び方")
            st.markdown("調査の目的に合わせて、最適な設定を選んでください。")
            
            # モード比較表
            comparison_data = {
                "モード設定": ["基本 (デフォルト)", "Proモードのみ", "Proモード + RDAP"],
                "こんな時におすすめ": ["とりあえず無料で調べたい", "正確な企業名・VPN判定が知りたい", "法的調査・徹底的に裏取りしたい"],
                "API設定": ["ip-api.com (無料)", "ipinfo.io (要APIキー)", "ipinfo.io (要APIキー)"],
                "オプション": ["RDAPオフ", "RDAPオフ", "RDAPオン ✅"],
                "処理速度": ["🚀 速い", "🚀 速い", "🐢 普通〜遅い"],
                "ISP情報の質": ["△ 古い (名寄せで補正)", "◯ 新しい (通称名)", "◎ 最も正確 (登記名)"],
                "VPN/Proxy判定": ["△ 推測 (名前ベース)", "◎ 正確 (DB照合)", "◎ 正確 (DB照合)"]
            }
            st.table(pd.DataFrame(comparison_data).set_index("モード設定"))
            
            st.markdown("""
            - **基本モード**: APIキー不要。名寄せ機能により、J:COMなどの主要ISPは正しく表示されます。
            - **Proモード**: サイドバーにAPIキーを入力すると有効化。VPNやHostingの判定精度が格段に上がります。
            - **RDAPオプション**: 「検索表示設定」でチェックを入れると有効化。公式台帳を参照し、データの完全性を高めます。
            """)

        with tab2:
            st.markdown("""
            #### 1. データソース
            - **IP Geolocation / ISP 情報**: 
                - 無料版: `ip-api.com` (毎分45リクエスト制限)
                - Pro版: `ipinfo.io` (APIキーに基づく制限)
            - **Whois (RDAP)**: APNIC等の各地域レジストリ公式サーバー
            - **Tor出口ノード**: Tor Project公式サイトより起動時に最新リストを取得

            #### 2. 強力な名寄せ機能
            `ip-api` 等のデータベースに残る古いISP名称（例: JCN, Jupiter, So-net等）を、独自辞書により現在のブランド名（例: J:COM, Sony, NTT等）に自動変換して集計します。これにより、表記揺れによる分析のストレスを軽減します。

            #### 3. 技術的仕様
            - **並列処理**: マルチスレッドによる高速検索（APIレートリミット自動調整機能付き）
            - **CIDRキャッシュ**: 同一ネットワーク帯域（/24など）への重複リクエストを回避し、高速化
            - **OCR誤読補正**: `1` と `l`、`0` と `O` などのOCR読み取りミスを自動修正して検索
            """)
            
            st.markdown("#### 4. 判定ステータスの意味")
            st.warning("⚠️ **Hosting/VPN/Proxy**")
            st.markdown("データセンター、VPNサービス、プロキシサーバー経由の通信です。一般家庭からのアクセスではなく、ボットや匿名化ツールを使用している可能性があります。")
            st.error("⚠️ **Tor Node**")
            st.markdown("Tor匿名化ネットワークの出口ノードです。攻撃の前兆や、高い匿名性を必要とする通信の可能性があります。")

        with tab3:
            st.markdown("""
            **Q. ファイルをアップロードしても大丈夫ですか？**
            A. このツールはローカル（またはセキュアなクラウド環境）で動作し、アップロードされたデータが外部（開発者）に送信されることはありません。APIへの問い合わせにはIPアドレスのみが送信されます。

            **Q. 検索が途中で止まりました。**
            A. APIの制限（レートリミット）にかかった可能性があります。ツールは自動的に待機して再開しますが、大量（数千件）の検索を行う場合は時間がかかります。「待機中」の表示が出ている場合はそのままお待ちください。

            **Q. ipinfoのAPIキーはどこで手に入りますか？**
            A. [ipinfo.io](https://ipinfo.io/signup) から無料で登録・取得できます（無料枠あり）。
            """)
        return
            

    # --- メインコンテンツ：Whois検索タブ ---
    st.title("🌐 検索大臣 - Whois & IP Intelligence -")

    col_input1, col_input2 = st.columns([1, 1])

    with col_input1:
        manual_input = st.text_area(
            "📋 テキスト入力 (IP/ドメイン)",
            height=150,
            placeholder="8.8.8.8\nexample.com\n2404:6800:..."
        )

    with col_input2:
        # --- モードによるアップロード制限の切り替え ---
        if IS_PUBLIC_MODE:
            # 公開モード (st版の挙動): txtのみ許可、警告あり
            allowed_types = ['txt']
            label_text = "📂 IPリストをアップロード (.txtのみ)"
            help_text = "※ 1行に1つのターゲットを記載"
        else:
            # ローカルモード (my版の挙動): csv/excel許可
            allowed_types = ['txt', 'csv', 'xlsx', 'xls']
            label_text = "📂 リストをアップロード (txt/csv/xlsx)"
            help_text = "※ 1行に1つのターゲットを記載、またはCSV/ExcelのIP列を自動検出します"

        uploaded_file = st.file_uploader(label_text, type=allowed_types)
        st.caption(help_text)
        
        raw_targets = []
        df_orig = None # 初期化

        if manual_input:
            raw_targets.extend(manual_input.splitlines())
        
        if uploaded_file:
            # --- 公開モードの場合の読み込み処理 (st版ロジック) ---
            if IS_PUBLIC_MODE:
                 try:
                    # シンプルにテキストとして読み込む
                    string_data = uploaded_file.read().decode("utf-8")
                    raw_targets.extend(string_data.splitlines())
                    
                    # 元データフレーム機能は無効化
                    st.session_state['original_df'] = None
                    st.session_state['ip_column_name'] = None
                    
                    st.info(f"📄 テキスト読み込み完了: {len(raw_targets)} 行")

                 except Exception as e:
                    st.error(f"ファイル読み込みエラー: {e}")
            
            # --- ローカルモードの場合の読み込み処理 (my版ロジック) ---
            else:
                ip_col = None
                try:
                    if uploaded_file.name.endswith('.csv'):
                        df_orig = pd.read_csv(uploaded_file)
                    elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                        df_orig = pd.read_excel(uploaded_file)
                    else:
                        # TXTファイル
                        raw_targets.extend(uploaded_file.read().decode("utf-8").splitlines())
                        st.session_state['original_df'] = None
                        st.session_state['ip_column_name'] = None

                    if df_orig is not None:
                        st.session_state['original_df'] = df_orig
                        for col in df_orig.columns:
                            sample = df_orig[col].dropna().head(10).astype(str)
                            if any(is_valid_ip(val.strip()) for val in sample):
                                ip_col = col
                                break
                        
                        if ip_col:
                            st.session_state['ip_column_name'] = ip_col
                            raw_targets.extend(df_orig[ip_col].dropna().astype(str).tolist())
                            
                            st.info(f"📄 ファイル読み込み完了: {len(df_orig)} 行 / IP列: `{ip_col}`")
                            with st.expander("👀 アップロードデータ・プレビュー", expanded=False):
                                st.dataframe(df_orig)
                        else:
                            st.error("ファイル内にIPアドレスの列が見つかりませんでした。")

                except Exception as e:
                    st.error(f"ファイル読み込みエラー: {e}")

    # --- 公開モード時のみセキュリティ警告を表示 ---
    if IS_PUBLIC_MODE:
        st.warning("""
        **🛡️ セキュリティ上の注意**
        * **テキスト入力推奨**: ファイルアップロードよりも、左側のテキストエリアへの**コピー＆ペースト**の方が、メタデータ（作成者情報など）が含まれないため安全です。
        * **ファイル名に注意**: アップロードする場合は、ファイル名に機密情報（例: `ClientA_Log.txt`）を含めず、`list.txt` などの無機質な名前を使用してください。
        """)
    
    cleaned_raw_targets_list = []
    target_freq_counts = {}

    if raw_targets:
        cleaned_raw_targets_list = [clean_ocr_error_chars(t) for t in raw_targets]
        target_freq_counts = pd.Series(cleaned_raw_targets_list).value_counts().to_dict()
    else:
        target_freq_counts = {}

    targets = []
    ocr_error_chars = set('Iil|OoSsAaBⅡ')

    for t in raw_targets:
        original_t = t
        is_ocr_error_likely = any(c in ocr_error_chars for c in original_t)
        if is_ocr_error_likely:
            cleaned_t = clean_ocr_error_chars(original_t)
            if is_valid_ip(cleaned_t):
                if cleaned_t not in targets: targets.append(cleaned_t)
                continue
            t = original_t
        
        invalid_ip_chars = set('ghijklmnopqrstuvwxyz')
        has_hyphen = '-' in t
        has_strictly_domain_char = any(c in invalid_ip_chars for c in t.lower())
        is_likely_domain_or_host = has_hyphen or has_strictly_domain_char
    
        if is_valid_ip(t):
            if t not in targets: targets.append(t)
        elif is_likely_domain_or_host:
            if t not in targets: targets.append(t)
        else:
            cleaned_t_final = clean_ocr_error_chars(t)
            if cleaned_t_final not in targets: targets.append(cleaned_t_final)

    has_new_targets = (targets != st.session_state.targets_cache)
    
    if has_new_targets or 'target_freq_map' not in st.session_state:
        st.session_state['target_freq_map'] = target_freq_counts
        st.session_state['original_input_list'] = cleaned_raw_targets_list
    ip_targets = [t for t in targets if is_valid_ip(t)]
    domain_targets = [t for t in targets if not is_valid_ip(t)]
    ipv6_count = sum(1 for t in ip_targets if not is_ipv4(t))
    ipv4_count = len(ip_targets) - ipv6_count

    st.markdown("---")
    st.markdown("### ⚙️ 検索表示設定")
    
    col_set1, col_set2 = st.columns(2)
    with col_set1:
        display_mode = st.radio(
            "**表示モード:** (検索結果の表示形式とAPI使用有無を設定)",
            ("標準モード", "集約モード (IPv4 Group)", "簡易モード (APIなし)"),
            key="display_mode_radio",
            horizontal=False
        )
    
    with col_set2:
        api_mode_selection = st.radio(
            "**API 処理モード:** (速度と安定性のトレードオフ)",
            list(MODE_SETTINGS.keys()),
            key="api_mode_radio",
            horizontal=False
        )
        # 🆕 RDAPオプション
        use_rdap_option = st.checkbox("🔍 高精度モード (RDAP公式台帳の併用 - 低速)", value=False, help="無料APIのISP情報に加え、RDAP(公式台帳)から最新のネットワーク名を取得します。通信が増えるため処理が遅くなります。")
    
    selected_settings = MODE_SETTINGS[api_mode_selection]
    max_workers = selected_settings["MAX_WORKERS"]
    delay_between_requests = selected_settings["DELAY_BETWEEN_REQUESTS"]
    rate_limit_wait_seconds = RATE_LIMIT_WAIT_SECONDS

    mode_mapping = {
        "標準モード": "標準モード (1ターゲット = 1行)",
        "集約モード (IPv4 Group)": "集約モード (IPv4アドレスをISP/国別でグループ化)",
        "簡易モード (APIなし)": "簡易モード (APIなし - セキュリティリンクのみ)"
    }
    current_mode_full_text = mode_mapping[display_mode]

    st.markdown("---")
    col_act1, col_act2 = st.columns([3, 1])

    is_currently_searching = st.session_state.is_searching and not st.session_state.cancel_search
    
    total_ip_targets_for_display = len(ip_targets) + len(st.session_state.deferred_ips)

    with col_act1:
        st.success(f"**Target:** IPv4: {ipv4_count} / IPv6: {ipv6_count} / Domain: {len(domain_targets)} (Pending: {len(st.session_state.deferred_ips)}) / **CIDR Cache:** {len(st.session_state.cidr_cache)}")
        if pro_api_key:
            st.info("🔑 **Pro Mode Active:** ipinfo.io データベースを使用します")

    with col_act2:
        if is_currently_searching:
            if st.button("❌ 中止", type="secondary", use_container_width=True):
                st.session_state.cancel_search = True
                st.session_state.is_searching = False
                st.session_state.deferred_ips = {}
                st.rerun()
        else:
            execute_search = st.button(
            "🚀 検索開始",
            type="primary",
            use_container_width=True,
            disabled=(len(targets) == 0 and len(st.session_state.deferred_ips) == 0)
            )

    if ('execute_search' in locals() and execute_search and (has_new_targets or len(st.session_state.deferred_ips) > 0)) or is_currently_searching:
        
        if ('execute_search' in locals() and execute_search and has_new_targets and len(targets) > 0):
            st.session_state.is_searching = True
            st.session_state.cancel_search = False
            st.session_state.raw_results = []
            st.session_state.deferred_ips = {}
            st.session_state.finished_ips = set()
            st.session_state.targets_cache = targets
            st.session_state.search_start_time = time.time()
            st.rerun() 
            
        elif is_currently_searching:
            targets = st.session_state.targets_cache
            ip_targets = [t for t in targets if is_valid_ip(t)]
            domain_targets = [t for t in targets if not is_valid_ip(t)]

            st.subheader("⏳ 処理中...")
            
            total_targets = len(targets)
            total_ip_api_targets = len(ip_targets)
            
            ip_targets_to_process = [ip for ip in ip_targets if ip not in st.session_state.finished_ips]
            
            current_time = time.time()
            ready_to_retry_ips = []
            deferred_ips_new = {}
            for ip, defer_time in st.session_state.deferred_ips.items():
                if current_time >= defer_time:
                    ready_to_retry_ips.append(ip)
                else:
                    deferred_ips_new[ip] = defer_time
            
            st.session_state.deferred_ips = deferred_ips_new
            
            immediate_ip_queue_unique = []
            for ip in ip_targets_to_process:
                if ip not in st.session_state.deferred_ips and ip not in immediate_ip_queue_unique:
                    immediate_ip_queue_unique.append(ip)

            immediate_ip_queue = immediate_ip_queue_unique
            immediate_ip_queue.extend(ready_to_retry_ips)
            
            if "簡易" in current_mode_full_text:
                if not st.session_state.raw_results:
                    results_list = []
                    for t in targets:
                        results_list.append(get_simple_mode_details(t))
                    st.session_state.raw_results = results_list
                    st.session_state.finished_ips.update(targets)
                    st.session_state.is_searching = False
                    st.rerun()

            else:
                if not any(res['ISP'] == 'Domain/Host' for res in st.session_state.raw_results) and domain_targets:
                    st.session_state.raw_results.extend([get_domain_details(d) for d in domain_targets])
                    st.session_state.finished_ips.update(domain_targets)
                    
                prog_bar_container = st.empty()
                status_text_container = st.empty()
                summary_container = st.empty() 

                if immediate_ip_queue:
                    cidr_cache_snapshot = st.session_state.cidr_cache.copy() 
                    
                    with ThreadPoolExecutor(max_workers=max_workers) as executor:
                        future_to_ip = {
                            executor.submit(
                                get_ip_details_from_api, 
                                ip, 
                                cidr_cache_snapshot, 
                                delay_between_requests, 
                                rate_limit_wait_seconds,
                                tor_nodes,
                                use_rdap_option,
                                pro_api_key # APIキーを渡す
                            ): ip for ip in immediate_ip_queue
                        }
                        remaining = set(future_to_ip.keys())
                        
                        while remaining and not st.session_state.cancel_search:
                            done, remaining = wait(remaining, timeout=0.1, return_when=FIRST_COMPLETED)
                            
                            for f in done:
                                res_tuple = f.result()
                                res = res_tuple[0]
                                new_cache_entry = res_tuple[1]
                                ip = res['Target_IP']
                                
                                if new_cache_entry:
                                    st.session_state.cidr_cache.update(new_cache_entry)
                                
                                if res.get('Status', '').startswith('Success'):
                                    st.session_state.raw_results.append(res)
                                    st.session_state.finished_ips.add(ip)
                                elif res.get('Defer_Until'):
                                    st.session_state.deferred_ips[ip] = res['Defer_Until']
                                else:
                                    st.session_state.raw_results.append(res)
                                    st.session_state.finished_ips.add(ip)

                            if total_ip_api_targets > 0:
                                processed_api_ips_count = len([ip for ip in st.session_state.finished_ips if is_valid_ip(ip)])
                                pct = int(processed_api_ips_count / total_ip_api_targets * 100)
                                elapsed_time = time.time() - st.session_state.search_start_time
                                eta_seconds = 0
                                if processed_api_ips_count > 0:
                                    rate = processed_api_ips_count / elapsed_time
                                    remaining_count = total_ip_api_targets - processed_api_ips_count
                                    eta_seconds = math.ceil(remaining_count / rate)
                                
                                eta_display = "計算中..."
                                if eta_seconds > 0:
                                    minutes = int(eta_seconds // 60)
                                    seconds = int(eta_seconds % 60)
                                    eta_display = f"{minutes:02d}:{seconds:02d}"
                                    
                                with prog_bar_container:
                                    st.progress(pct)
                                with status_text_container:
                                    st.caption(f"**Progress:** {processed_api_ips_count}/{total_ip_api_targets} | **Deferred:** {len(st.session_state.deferred_ips)} | **CIDR Cache:** {len(st.session_state.cidr_cache)} | **Remaining Time:** {eta_display}")
                                
                                isp_df, country_df, freq_df, country_all_df, isp_full_df, country_full_df, freq_full_df = summarize_in_realtime(st.session_state.raw_results)
                                with summary_container.container():
                                    st.markdown("---")
                                    draw_summary_content(isp_df, country_df, freq_df, country_all_df, "📊 Real-time analysis")
                                st.markdown("---")

                            if not remaining and not st.session_state.deferred_ips:
                                break
                            
                            if st.session_state.deferred_ips:
                                st.rerun()  
                            
                            time.sleep(0.5) 
                            
                        if total_ip_api_targets > 0 and not st.session_state.deferred_ips:
                            processed_api_ips_count = len([ip for ip in st.session_state.finished_ips if is_valid_ip(ip)])
                            final_pct = int(processed_api_ips_count / total_ip_api_targets * 100)
                            with prog_bar_container:
                                st.progress(final_pct)
                            with status_text_container:
                                st.caption(f"**Progress:** {processed_api_ips_count}/{total_ip_api_targets} | **Deferred:** {len(st.session_state.deferred_ips)} | **CIDR Cache:** {len(st.session_state.cidr_cache)} | **Remaining Time:** 完了")
                        
                if len(st.session_state.finished_ips) == total_targets and not st.session_state.deferred_ips:
                    st.session_state.is_searching = False
                    st.info("✅ 全ての検索が完了しました。")
                    summary_container.empty()
                    st.rerun()
                
                elif st.session_state.deferred_ips and not st.session_state.cancel_search:
                    next_retry_time = min(st.session_state.deferred_ips.values())
                    wait_time = max(1, int(next_retry_time - time.time()))
                    
                    prog_bar_container.empty()
                    status_text_container.empty()
                    summary_container.empty()
                    st.warning(f"⚠️ **APIレートリミットに達しました。** 隔離中の **{len(st.session_state.deferred_ips)}** 件のIPアドレスは **{wait_time}** 秒後に再試行されます。")
                    time.sleep(min(5, wait_time)) 
                    st.rerun()

                elif st.session_state.cancel_search:
                    prog_bar_container.empty()
                    status_text_container.empty()
                    summary_container.empty()
                    st.warning("検索がユーザーによって中止されました。")
                    st.session_state.is_searching = False
                    st.rerun()


    # --- 結果表示 ---
    if st.session_state.raw_results or st.session_state.deferred_ips:
        res = st.session_state.raw_results
        
        if st.session_state.get('debug_summary'):
            with st.expander("🛠️ デバッグ情報 (集計データ確認用)", expanded=False):
                st.markdown("**API 処理モード設定**")
                st.write(f"MAX_WORKERS: {max_workers}")
                st.write(f"DELAY_BETWEEN_REQUESTS: {delay_between_requests}")
                st.markdown("---")
                st.json(st.session_state['debug_summary'].get('country_code_counts', {}))
                st.json(st.session_state['debug_summary'].get('country_all_df', []))
                st.markdown("---")
                st.json(st.session_state.get('cidr_cache', {}))

        
        successful_results = [r for r in res if r['Status'].startswith('Success') or r['Status'].startswith('Aggregated')]
        error_results = [r for r in res if not (r['Status'].startswith('Success') or r['Status'].startswith('Aggregated'))]
        
        for ip, defer_time in st.session_state.deferred_ips.items():
            status = f"Pending (Retry in {max(0, int(defer_time - time.time()))}s)"
            error_results.append({
                'Target_IP': ip, 'ISP': 'N/A', 'Country': 'N/A', 'CountryCode': 'N/A', 'RIR_Link': get_authoritative_rir_link(ip, 'N/A'),
                'Secondary_Security_Links': create_secondary_links(ip), 
                'Status': status
            })
        
        if "集約" in current_mode_full_text:
            display_res = group_results_by_isp(successful_results)
            display_res.extend(error_results)
        else:
            display_res = successful_results + error_results
            target_order = {ip: i for i, ip in enumerate(targets)}
            display_res.sort(key=lambda x: target_order.get(get_copy_target(x['Target_IP']), float('inf')))

        display_results(display_res, current_mode_full_text, display_mode)
        
        if not st.session_state.is_searching or st.session_state.cancel_search:
            isp_df, country_df, freq_df, country_all_df, isp_full_df, country_full_df, freq_full_df = summarize_in_realtime(st.session_state.raw_results)
            
            st.markdown("---")
            draw_summary_content(isp_df, country_df, freq_df, country_all_df, "✅ 集計結果")

            # --- 元データ結合処理（画面表示 & ダウンロード共通） ---
            df_with_res = pd.DataFrame() # 初期化
            if st.session_state.get('original_df') is not None and st.session_state.get('ip_column_name'):
                df_with_res = st.session_state['original_df'].copy()
                ip_col = st.session_state['ip_column_name']
                results = st.session_state.get('raw_results', []) 
                
                if results:
                    res_dict = {r['Target_IP']: r for r in results}

                    # 各行のIPに基づいて結果をマッピング
                    isps, isps_jp, countries, countries_jp, proxy_type, statuses = [], [], [], [], [], []
                    for ip_val in df_with_res[ip_col]:
                        ip_val_str = str(ip_val).strip()
                        info = res_dict.get(ip_val_str, {})
                        isps.append(info.get('ISP', 'N/A'))
                        isps_jp.append(info.get('ISP_JP', 'N/A')) 
                        countries.append(info.get('Country', 'N/A'))
                        countries_jp.append(info.get('Country_JP', 'N/A'))
                        proxy_type.append(info.get('Proxy_Type', ''))
                        statuses.append(info.get('Status', 'N/A'))
                    
                    # 結合
                    insert_idx = df_with_res.columns.get_loc(ip_col) + 1
                    df_with_res.insert(insert_idx, 'Status', statuses)
                    df_with_res.insert(insert_idx, 'Proxy Type', proxy_type)
                    df_with_res.insert(insert_idx, 'Country_JP', countries_jp)
                    df_with_res.insert(insert_idx, 'Country', countries)
                    df_with_res.insert(insert_idx, 'ISP_JP', isps_jp)
                    df_with_res.insert(insert_idx, 'ISP', isps)

            # --- 新機能：元データ x 検索結果 クロス分析表示 ---
            if not df_with_res.empty:
                st.markdown("---")
                render_merged_analysis(df_with_res)
            # ------------------------------------------------

            # --- 全件集計データのダウンロードセクション ---
            st.markdown("### 📊 集計データの完全版ダウンロード")
            st.caption("※ 上記グラフのTop10制限を解除した、すべての集計データとグラフをダウンロードできます。")
            
            col_full_dl1, col_full_dl2, col_full_dl3, col_full_dl4 = st.columns(4)
            
            with col_full_dl1:
                st.download_button(
                    "⬇️ 対象IP カウント (全件)",
                    freq_full_df.to_csv(index=False).encode('utf-8-sig'),
                    "target_ip_frequency_all.csv",
                    "text/csv",
                    use_container_width=True
                )
            with col_full_dl2:
                st.download_button(
                    "⬇️ ISP別 カウント (全件)",
                    isp_full_df.to_csv(index=False).encode('utf-8-sig'),
                    "isp_counts_all.csv",
                    "text/csv",
                    use_container_width=True
                )
            with col_full_dl3:
                st.download_button(
                    "⬇️ 国別 カウント (全件)",
                    country_full_df.to_csv(index=False).encode('utf-8-sig'),
                    "country_counts_all.csv",
                    "text/csv",
                    use_container_width=True
                )
            
            with col_full_dl4:
                # 全件グラフHTMLレポートの生成
                html_report = generate_full_report_html(isp_full_df, country_full_df, freq_full_df)
                st.download_button(
                    "⬇️ 全件グラフHTMLレポート",
                    html_report,
                    "whois_analysis_report.html",
                    "text/html",
                    use_container_width=True
                )

        
        st.markdown("### ⬇️ 検索結果リストのダウンロード")
        col_dl1, col_dl2, col_dl3 = st.columns(3)
        # 1. 画面表示順データ
        csv_display = pd.DataFrame(display_res).drop(columns=['CountryCode', 'Secondary_Security_Links', 'RIR_Link'], errors='ignore').astype(str)
        with col_dl1:
            st.download_button("⬇️ CSV (画面表示順)", csv_display.to_csv(index=False).encode('utf-8-sig'), "whois_results_display.csv", "text/csv", use_container_width=True)
            # Excel (Display)
            excel_display = convert_df_to_excel(csv_display)
            st.download_button("⬇️ Excel (画面表示順)", excel_display, "whois_results_display.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        # 2. 全入力データ（入力順）
        result_lookup = {r['Target_IP']: r for r in st.session_state.raw_results}
        full_output_data = []
        for original_t in st.session_state.get('original_input_list', []):
            if original_t in result_lookup:
                full_output_data.append(result_lookup[original_t])
            else:
                full_output_data.append({'Target_IP': original_t, 'ISP': 'N/A', 'ISP_JP': 'N/A', 'Country': 'N/A', 'Country_JP': 'N/A', 'Status': 'Pending/Error'})
        
        csv_full = pd.DataFrame(full_output_data).drop(columns=['CountryCode', 'Secondary_Security_Links', 'RIR_Link'], errors='ignore').astype(str)
        with col_dl2:
            st.download_button("⬇️ CSV (全入力データ順)", csv_full.to_csv(index=False).encode('utf-8-sig'), "whois_results_full.csv", "text/csv", use_container_width=True)
            # Excel (Full)
            excel_full = convert_df_to_excel(csv_full)
            st.download_button("⬇️ Excel (全入力データ順)", excel_full, "whois_results_full.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        with col_dl3:
            # 3. 元データ結合ダウンロード（共通処理で作成済みのdf_with_resを使用）
            if not IS_PUBLIC_MODE and not df_with_res.empty:
                st.markdown("**🔍 分析付きExcel (Pivot/Graph)**")
                
                # 時間帯分析用の列選択ボックス
                time_cols = [c for c in df_with_res.columns if 'date' in c.lower() or 'time' in c.lower() or 'jst' in c.lower()]
                default_idx = df_with_res.columns.get_loc(time_cols[0]) if time_cols else 0
                
                selected_time_col = st.selectbox(
                    "時間帯分析(Hour列)に使う日時列を選択:", 
                    df_with_res.columns, 
                    index=default_idx,
                    key="time_col_selector"
                )

                # Advanced Excel生成 (v5.0)
                excel_advanced = create_advanced_excel(df_with_res, selected_time_col)
                
                st.download_button(
                    "⬇️ Excel (分析・グラフ付き)", 
                    excel_advanced, 
                    "whois_analysis_master.xlsx", 
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                    use_container_width=True,
                    help="生データに加え、ISP別・時間帯別の集計表とグラフ（ピボット）が別シートに含まれます。"
                )
            else:
                st.button("⬇️ Excel (CSVアップロード時のみ)", disabled=True, use_container_width=True)

if __name__ == "__main__":
    main()
